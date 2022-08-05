import openpyxl
import pymysql.cursors
import os
from dotenv import dotenv_values

sheet_names = ['Hoja1']
file = "ACTUALIZAR PVP.xlsx"
config = {
    **dotenv_values("db.env")
}

def PrintDimensions(sheet):
    print("Dimensiones -> ", sheet.dimensions)

def PrintRowsAndColumns(sheet):
    print("MIN ROW {} - MAX ROW {}".format(sheet.min_row, sheet.max_row))
    print("MIN COLUMN {} - MAX COLUMN {}".format(sheet.min_column, sheet.max_column))

def PrintData(sheet):
    for row in sheet.rows:
        print(*[cell.value for cell in row])

def test_db():
    connection = pymysql.connect(host=config['host'], user=config['user'], password=config['password'], database=config['database'])
    with connection:
        with connection.cursor() as cursor:
            sql = "select @@version"
            cursor.execute(sql)
            result = cursor.fetchone()
            print(result)

def main():
    sqlStatement = ''
    wb = openpyxl.load_workbook(file)
    # type(wb) -> <class 'openpyxl.workbook.workbook.Workbook'>
    try:
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            PrintDimensions(sheet)
            PrintRowsAndColumns(sheet)
            for row in range(2, sheet.max_row + 1):
                sqlStatement = 'UPDATE producto_producto SET precio = {} WHERE codigo = "{}"'.format(round(sheet['C' + str(row)].value, 2), sheet['A' + str(row)].value)
                print(sqlStatement)
    except KeyError as e:
        print("No se encontró la hoja en el Worksheet")

if __name__ == "__main__":
    test_db()
    # main()
